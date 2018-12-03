import os
import random
import re
import socket
import threading
import time
import urllib
from urllib import request

import emoji
import jieba
import matplotlib.pyplot as plt
import openpyxl
import requests
from PIL import Image
from aip import AipNlp
from bs4 import BeautifulSoup
from pylab import mpl
from wordcloud import WordCloud

# 代理开关
using_proxy = 0
# 情感分析开关
emotion_ai = 1

ip_pool = []
proxies_list = []
rank_list = []
now_movie_list = []
emotion_list = []
jieba.add_word('任素汐')

# Baidu_Aip密钥
APP_ID = '14963270'
API_KEY = 'DzcClKytSKKGpbTxdSFUcyif'
SECRET_KEY = 'IIIiZbKyHdFDqA7xq17kb8wS3elRVF7a'
# 加载百度Ai自然语言处理
client = AipNlp(APP_ID, API_KEY, SECRET_KEY)
# http_headers
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64)\
   AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2272.118 Safari/537.36', }


# 获取代理的IP_pool并多线程验证,写入文件
def get_proxies():
    for page in range(1, 5):
        IPurl = 'http://www.xicidaili.com/nt/%s' % page
        rIP = requests.get(IPurl, headers=headers)
        IPContent = rIP.text
        soupIP = BeautifulSoup(IPContent, "html.parser")
        trs = soupIP.find_all('tr')
        for tr in trs[1:]:
            tds = tr.find_all('td')
            ip = tds[1].text.strip()
            port = tds[2].text.strip()
            protocol = tds[5].text.strip()
            if protocol == 'HTTP':
                # 由于我们爬取的网站为HTTPS所以并不需要HTTP的代理
                httpResult = ip + ':' + port
            elif protocol == 'HTTPS':
                httpsResult = ip + ':' + port
                ip_pool.append({"https": httpsResult})

    proxy_ip = open('proxy_ip.txt', 'w')  # 新建一个储存有效IP的文档
    lock = threading.Lock()  # 建立一个锁

    def test(i):
        socket.setdefaulttimeout(5)  # 设置全局超时时间
        url = "https://movie.douban.com/"
        try:
            proxy_support = urllib.request.ProxyHandler(ip_pool[i])
            opener = urllib.request.build_opener(proxy_support)
            opener.addheaders = [("User-Agent", "Mozilla/5.0 (Windows NT 10.0; WOW64)")]
            urllib.request.install_opener(opener)
            res = urllib.request.urlopen(url).read()
            lock.acquire()  # 获得锁
            print(ip_pool[i], 'is OK')
            proxy_ip.write('%s\n' % str(ip_pool[i]))  # 写入该代理IP
            proxies_list.append(ip_pool[i])
            lock.release()  # 释放锁
        except Exception as e:
            lock.acquire()
            print(ip_pool[i], e)
            lock.release()

    # 多线程验证
    threads = []
    for i in range(len(ip_pool)):
        thread = threading.Thread(target=test, args=[i])
        threads.append(thread)
        thread.start()
    # 阻塞主进程，等待所有子线程结束
    for thread in threads:
        thread.join()

    proxy_ip.close()  # 关闭文件


# 初始化代理服务器（取proxy_list[i]）
def proxy_initialize(i):
    proxy_support = urllib.request.ProxyHandler(proxies_list[i])
    opener = urllib.request.build_opener(proxy_support)
    opener.addheaders = [("User-Agent", "Mozilla/5.0 (Windows NT 10.0; WOW64)")]
    urllib.request.install_opener(opener)


# 通过代理连接ip.cn返回当前IP地址以显示代理的可用性
def proxy_verify(i):
    proxy_initialize(i)
    print(proxies_list[i])
    res = urllib.request.urlopen('https://www.ip.cn').read()
    print(re.findall(r"您(.+?)China", res.decode()))


# 模拟登录
def login_douban(redir):
    if using_proxy:
        proxy_initialize(0)
    login_url = "https://accounts.douban.com/login"
    my_post = {'redir': redir,
               'form_email': '1505466304@qq.com',
               'form_password': '6305318ST',
               'login': '登录',
               }
    r = requests.post(login_url, data=my_post, headers=headers)
    html = r.text
    if html.__contains__('try'):
        print(html)

    # 如果直接登录成功则返回
    if r.url == redir:
        return html
    # 否则拉取captcha
    reg = r'<img id="captcha_image" src="(.*?)" alt="captcha" class="captcha_image"/>'
    imglist = re.findall(reg, html)
    pic = 'captchas\\%d.jpg' % random.randint(1, 100)
    urllib.request.urlretrieve(imglist[0], pic)
    img = Image.open(pic)
    # 调用plt在内部绘制captcha
    plt.figure("captcha")
    plt.imshow(img)
    plt.show()
    # 手动输入captcha
    captcha = input('captcha is: ')
    regid = r'<input type="hidden" name="captcha-id" value="(.*?)"/>'
    ids = re.findall(regid, html)
    my_post["captcha-solution"] = captcha
    my_post["captcha-id"] = ids[0]
    q = requests.post(login_url, data=my_post, headers=headers)
    print(q.url)
    return q.text


# jieba分词
def jieba_split(txt):
    text = open(txt, "rb").read()
    words_list = []
    word_generator = jieba.cut(text, cut_all=False)
    # 加载停止词并去除
    with open('chineseStopWords.txt') as f:
        str_text = f.read()
        f.close()
    for word in word_generator:
        if word.strip() not in str_text:
            words_list.append(word)
    return ' '.join(words_list)


# 调用api分析评论情感
def get_sentiments(text):
    try:
        sitems = client.sentimentClassify(text)['items'][0]  # 情感分析
        print(sitems)
        print('\n')
        # 积极概率 positive = sitems['positive_prob']
        # 置信度 confidence = sitems['confidence']
        # 0表示消极，1表示中性，2表示积极 sentiment = sitems['sentiment']
        return sitems
    except Exception as e:
        print(e)


# 获取正在热映电影
def get_now_movie(html):
    soup = BeautifulSoup(html)
    div_list = soup.body.find(id='screening')
    now_movie_taglist = div_list.find_all('li', class_=re.compile(r'^ui-slide-item\s?s?'))

    for item in now_movie_taglist:
        now_movie_dict = {'name': item.get('data-title'), 'id': item.get('data-trailer'),
                          'score': item.get('data-rate')}
        if now_movie_dict['name'] is None:
            pass
        else:
            now_movie_list.append(now_movie_dict)

    for movie in now_movie_list:
        movie['name'] = re.findall(r'\w+', movie['name'])[0]
        movie['id'] = re.findall(r'\d+', movie['id'])[0]
        print(movie['name'] + '_' + movie['score'] + '分')


# 根据短评页面url获取评论
def getcomment(comment_url_list, filename, switch):
    comments_list = []
    n = 0
    sum = 0
    count = 0
    count_e = 0
    for url in comment_url_list:
        if using_proxy:
            while 1:
                try:
                    proxy_initialize(count // 10 + count_e)
                    print('using proxy:', proxies_list[count // 10 + count_e])
                    comment_res = urllib.request.urlopen(url)
                    print(url)
                    count = count + 1
                    break
                except Exception as e:
                    print(e)
                    count_e = count_e + 1
        else:
            comment_res = urllib.request.urlopen(url)

        comment_soup = BeautifulSoup(comment_res, 'html.parser')
        comments = [item.string for item in comment_soup.find_all('span', class_='short')]
        comments_list.append(comments)
        if switch:
            # 导出评论及评分list
            tags = [item for item in comment_soup.findAll('div', class_='comment')]
            try:
                for i in range(20):
                    rank = {
                        'comments': comments[i],
                        # 使用正则表达式提取评分
                        'rank': re.findall(r"allstar(.+?)0 rating", str(tags[i])),
                    }
                    # 只收集评分不为3且长度大于4的评论
                    if rank['rank'] and int(rank['rank'][0]) != 3 and len(rank['comments']) > 4:
                        print(rank)
                        rank_list.append(rank)
            except Exception as e:
                print(e)

    if not switch:
        # 将分词结果写入文件
        with open('comments\\' + filename + '.txt', 'w', encoding='utf-8') as file:
            print('\n' + filename + '\n')
            for com_list in comments_list:
                for com in com_list:
                    print(com)
                    # 过滤标点符号
                    pattern = re.compile(r'[\u4e00-\u9fa5]+')
                    filterdata = re.findall(pattern, str(com))
                    cleaned_comments = ''.join(filterdata)
                    file.write(cleaned_comments + '\n')

                    if emotion_ai:
                        # 获取情感分析并计算加权平均
                        sitems = get_sentiments(emoji.demojize(com))
                        try:
                            sum = sum + sitems['positive_prob'] * sitems['confidence']
                            n = n + sitems['confidence']
                        except Exception as e:
                            print(e)

        if emotion_ai:
            emotion = {'movie': filename,
                       'emotion': round(sum / n, 2),
                       }
            print(emotion)
            emotion_list.append(emotion)

        create_wordcloud(jieba_split('comments\\' + filename + '.txt'), filename)


# 生成词云
def create_wordcloud(wl, filename):
    wc = WordCloud(background_color="white",
                   max_words=2000,
                   scale=16,
                   font_path="C:\Windows\Fonts\sourcehansans.ttf",
                   max_font_size=60,
                   random_state=30,
                   )
    try:
        myword = wc.generate(wl)
        wc.to_file('wordclouds\\' + filename + '.jpg')
        # 自定义图片标题字体以显示中文
        mpl.rcParams['font.sans-serif'] = ['SimHei']
        plt.imshow(myword)
        plt.title(filename)
        plt.axis("off")
        plt.show()
    except BaseException as e:
        print(e)


# 导出评论及星级到excel
def ex_ranks():
    comment_url_list = []
    # 自定义爬取电影id
    custom_id = ['26366496', '26752088', '27102569',
                 '26363254', '20376577', '26850326',
                 '27605698', '26985127', '26996640',
                 '26698897', '25823277', '25911694',
                 '26667056', '26753020', '1292213',
                 '3742360', '26861685', '4920528',
                 '26683723', '26575103', '30122633']
    for i in range(len(custom_id)):
        for page in range(11):
            comment_url_list.append('https://movie.douban.com/subject/' + custom_id[i] + '/comments?start=' + str(
                20 * page) + '&limit=20&sort=new_score&status=P')
        getcomment(comment_url_list, 'custom', 1)
        comment_url_list = []

    # 导出短评及评分到excel
    excel = openpyxl.Workbook()
    ws1 = excel.worksheets[0]
    ws1.title = u"Sheet1"
    ws1.cell(row=1, column=1).value = 'comment'
    ws1.cell(row=1, column=2).value = 'star'
    for i in range(len(rank_list)):
        ws1.cell(row=i + 2, column=1).value = emoji.demojize(rank_list[i]['comments'])
        ws1.cell(row=i + 2, column=2).value = int(rank_list[i]['rank'][0])
    excel.save("ranks.xlsx")


if __name__ == '__main__':
    # 建立comments, wordclouds, captchas文件夹以存放评论、词云和验证码
    if not os.path.exists('comments\\'):
        os.makedirs('comments\\')
    if not os.path.exists('wordclouds\\'):
        os.makedirs('wordclouds\\')
    if not os.path.exists('captchas\\'):
        os.makedirs('captchas\\')

    if using_proxy:
        get_proxies()

    # 通过代理连接ip.cn以显示其可用性
    # proxy_verify(i)

    while True:
        html = login_douban('https://movie.douban.com')
        time.sleep(1)
        if BeautifulSoup(html).title.get_text() == '登录豆瓣':
            print('登录失败，请等待3秒')
            time.sleep(3)
        else:
            print('登录成功')
            break

    # 导出评星到excel，使用时去掉注释
    # ex_ranks()

    # 获取正在热映的电影list
    get_now_movie(html)

    # 添加短评每页的url到comment_url_list
    comment_url_list = []
    for item in now_movie_list:
        for page in range(11):
            comment_url_list.append('https://movie.douban.com/subject/' + item['id'] + '/comments?start=' + str(
                20 * page) + '&limit=20&sort=new_score&status=P')
        getcomment(comment_url_list, item['name'] + '_' + item['score'] + '分', 0)
        comment_url_list = []

    # 打印每部电影的情感分析评分（0为负面，1为正面）
    if emotion_ai:
        for i in emotion_list:
            print(i)